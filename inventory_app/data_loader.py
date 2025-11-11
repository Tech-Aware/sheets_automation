"""Utilities responsible for loading the Excel workbook used by the UI."""
from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook


# ``dataclasses`` gained ``slots`` support in Python 3.10.  Older runtimes, such as
# the Python 3.9 interpreter bundled with some Windows installations, do not accept
# the argument which previously caused the application to fail at import time.
_DATACLASS_KWARGS = {"slots": True} if sys.version_info >= (3, 10) else {}


@dataclass(**_DATACLASS_KWARGS)
class SheetData:
    """Container describing the contents of a worksheet."""

    name: str
    headers: List[str]
    rows: List[Dict[str, Any]]

    def as_column_major(self) -> Dict[str, List[Any]]:
        """Return the sheet as a mapping of header -> column values."""

        return {header: [row.get(header) for row in self.rows] for header in self.headers}


class WorkbookRepository:
    """Loads data from the accounting workbook on demand."""

    def __init__(self, workbook_path: Path | str) -> None:
        self._workbook_path = Path(workbook_path)
        if not self._workbook_path.exists():
            raise FileNotFoundError(self._workbook_path)
        self._cache: Dict[str, SheetData] = {}

    @property
    def workbook_path(self) -> Path:
        return self._workbook_path

    def sheet_names(self) -> List[str]:
        """Return the sheet names available in the workbook."""

        with load_workbook(self._workbook_path, read_only=True, data_only=True) as wb:
            return list(wb.sheetnames)

    def refresh(self) -> None:
        """Invalidate the cached sheets."""

        self._cache.clear()

    def get_sheet(self, sheet_name: str, headers: Optional[Iterable[str]] = None) -> SheetData:
        """Return ``SheetData`` for ``sheet_name``.

        Parameters
        ----------
        sheet_name:
            Name of the sheet to load.
        headers:
            Optional ordered list of headers to prefer. When provided the loader will
            normalise the header casing and fill missing values with ``None``.
        """

        normalised_name = sheet_name.strip()
        if normalised_name in self._cache:
            return self._cache[normalised_name]

        with load_workbook(self._workbook_path, read_only=True, data_only=True) as wb:
            if normalised_name not in wb.sheetnames:
                raise KeyError(f"Sheet '{sheet_name}' not found in workbook {self._workbook_path!s}")
            ws = wb[normalised_name]

            iter_rows = ws.iter_rows(values_only=True)
            try:
                first_row = next(iter_rows)
            except StopIteration as exc:  # pragma: no cover - defensive guard
                raise ValueError(f"Sheet '{sheet_name}' is empty") from exc

            inferred_headers = [str(value).strip() if value is not None else f"COL_{idx+1}" for idx, value in enumerate(first_row)]
            header_lookup = _build_header_lookup(inferred_headers, headers)

            rows: List[Dict[str, Any]] = []
            for raw_row in iter_rows:
                row_dict = {header: raw_row[col_idx] if col_idx < len(raw_row) else None for header, col_idx in header_lookup.items()}
                if all(value in (None, "") for value in row_dict.values()):
                    continue
                rows.append(row_dict)

        sheet = SheetData(normalised_name, list(header_lookup.keys()), rows)
        self._cache[normalised_name] = sheet
        return sheet


def _build_header_lookup(inferred_headers: List[str], preferred_headers: Optional[Iterable[str]]) -> Dict[str, int]:
    """Create a mapping of header -> column index.

    The function is resilient to casing differences and missing headers. Preferred
    headers are returned in the order supplied; unspecified headers fall back to the
    inferred names gathered from the worksheet.
    """

    header_map: Dict[str, int] = {}
    canonical = {header.lower(): idx for idx, header in enumerate(inferred_headers)}

    if preferred_headers:
        for header in preferred_headers:
            idx = canonical.get(header.lower())
            if idx is not None:
                header_map[header] = idx
            else:
                header_map[header] = len(header_map)

    for idx, header in enumerate(inferred_headers):
        if header not in header_map:
            header_map[header] = idx

    return header_map


def load_named_sheets(repository: WorkbookRepository, requested: Dict[str, Iterable[str]]) -> Dict[str, SheetData]:
    """Batch load ``requested`` sheets.

    Parameters
    ----------
    repository:
        Source repository responsible for loading workbook data.
    requested:
        Mapping of sheet name -> iterable of preferred headers. If an iterable is
        empty the loader falls back to worksheet headers.
    """

    loaded: Dict[str, SheetData] = {}
    for sheet_name, headers in requested.items():
        loaded[sheet_name] = repository.get_sheet(sheet_name, headers=headers)
    return loaded
