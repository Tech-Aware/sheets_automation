"""Utilities to load the Excel source workbook."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, List, Sequence
import sys

try:  # pragma: no cover - optional dependency detection
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - fallback when optional dep missing
    load_workbook = None  # type: ignore


# ``dataclasses`` added support for ``slots`` in Python 3.10.  The production
# environment still runs Python 3.9, so silently drop the argument there while
# keeping the optimization available on newer interpreters.
_DATACLASS_KWARGS = {"slots": True} if sys.version_info >= (3, 10) else {}


@dataclass(**_DATACLASS_KWARGS)
class TableData:
    """Represents a sheet: its headers and a list of rows as dicts."""

    headers: Sequence[str]
    rows: List[dict]

    def head(self, count: int = 5) -> List[dict]:
        return self.rows[:count]


class WorkbookRepository:
    """Load the Excel workbook once and provide structured tables."""

    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():  # pragma: no cover - IO guard
            raise FileNotFoundError(self.workbook_path)
        self._workbook = None

    def _ensure_workbook(self):
        if self._workbook is None:
            if load_workbook is None:  # pragma: no cover - dependency guard
                raise RuntimeError(
                    "openpyxl is required to read the Excel workbook. Install it via 'pip install openpyxl'."
                )
            self._workbook = load_workbook(self.workbook_path, data_only=True)
        return self._workbook

    def available_tables(self) -> Iterable[str]:
        wb = self._ensure_workbook()
        return wb.sheetnames

    def load_table(self, sheet_name: str, *, drop_empty: bool = True) -> TableData:
        wb = self._ensure_workbook()
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found in workbook {self.workbook_path!s}")
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            return TableData(headers=[], rows=[])
        headers = [str(cell).strip() if cell is not None else "" for cell in raw_rows[0]]
        rows: List[dict] = []
        for raw in raw_rows[1:]:
            if drop_empty and all(value in (None, "") for value in raw):
                continue
            row_dict = {headers[idx]: self._normalize_value(raw[idx]) for idx in range(len(headers))}
            rows.append(row_dict)
        return TableData(headers=headers, rows=rows)

    def load_many(self, *sheet_names: str) -> dict[str, TableData]:
        return {name: self.load_table(name) for name in sheet_names}

    @staticmethod
    def _normalize_value(value):
        """Convert Excel date/time values to user-friendly strings."""

        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return value


__all__ = ["WorkbookRepository", "TableData"]
